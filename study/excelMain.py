# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import pandas as pd
import win32com.client as win32
import win32gui
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

which_menu = ""
k5_menu = ""
k4_menu = ""


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def webcrawling():
    URL = 'https://intranet.amkor.co.kr/'

    # URL = 'https://news.naver.com//'
    driver = webdriver.Chrome(executable_path='chromedriver')
    print(driver.get(url=URL))
    element = driver.get(url=URL)

    element = driver.find_element_by_xpath("//*")
    print(element)
    tagname = driver.find_elements_by_tag_name("iframe")
    print(tagname)
    driver.switch_to.frame(tagname[0])

    try:
        # element = WebDriverWait(driver, 5).until(
        #     EC.presence_of_element_located((By.CLASS_NAME , 'home_timelate'))
        # )
        # print(element)

        # elc=driver.find_element_by_xpath("/ html / body / div / div[6] / div[1] / div / div[7] / div[2] / div / ul / li[1] / a")
        # print(elc)

        # WebDriverWait(driver,5).until(EC.presence_of_element_located((By.XPATH,"/ html / body / div / div[3]/ div[3] / div / div / form / fieldset / input[1]")))
        #
        # searchbox=driver.find_element_by_xpath("/ html / body / div / div[3]/ div[3] / div / div / form / fieldset / input[1]")
        #
        # searchbox.send_keys("화이자")
        # searchbox.send_keys(Keys.ENTER)

        # k5 menu 읽기
        print("★☆★☆★☆★☆K5 Menu 입니다 ☆★☆★☆★☆★")
        xpathlink = "//*[@id=\"div-tabpanel-content-K5\"]"
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpathlink)))

        lunchmenu = driver.find_element_by_xpath(xpathlink)
        print(lunchmenu.text)

        k5_menu = lunchmenu.text

        print("=============================")
        # k4 menu 읽기
        print("★☆★☆★☆★☆K4 Menu 입니다 ☆★☆★☆★☆★")
        k4_xpath_link = "//*[@id=\"div-tabpanel-content-K4\"]"

        k4_menu_button_path = "//*[@id=\"tab-1012-btnEl\"]"

        k4_menu_button = driver.find_element_by_xpath(k4_menu_button_path)
        k4_menu_button.click()
        driver.implicitly_wait(1)

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, k4_xpath_link)))

        lunchmenu = driver.find_element_by_xpath(k4_xpath_link)
        print(lunchmenu.text)
        k4_menu = lunchmenu.text

        # popuppage="/html/body/div/img"
        # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, popuppage)))
        # element= driver.find_element_by_xpath(popuppage)
        # print(element)

        #
        # searchbox.send_keys("점심 메뉴")
        # searchbox.send_keys(Keys.ENTER)

        # sleep(5)

    finally:
        driver.quit()
        return k5_menu, k4_menu


# Press the green button in the gutter to run the script.
k4_menu_table = []
k5_menu_table = []


def create_excel_of_lunch_menu():
    global k5_menu, k4_menu;

    k5_menu, k4_menu = webcrawling()

    write_wb = Workbook()

    # 이름이 있는 시트를 생성
    # write_ws = write_wb.create_sheet('생성시트')

    # Sheet1에다 입력

    print("★☆★☆★☆★☆K5 Menu 입니다 ☆★☆★☆★☆★")
    row_idx = 2
    col_idx = 1

    write_ws = write_wb.active
    # write_ws = write_wb.create_sheet('k5 lunch menu')
    write_ws.title = "k5 menu"

    global which_menu
    which_menu = ""

    for i in k5_menu.split('\n'):
        if "중식" in i or "조식" in i or "석식" in i:
            write_ws['A1'] = i.split(" ")[0]
            which_menu = i.split(" ")[0]

            for j in i.split(" "):
                if "한식" in j:
                    row_idx += 2
                    col_idx = 1
                    write_ws.cell(row_idx, col_idx, j)
                    col_idx += 1
                    row_idx += 1
                elif which_menu not in j:
                    write_ws.cell(row_idx, col_idx, j)


        elif "일품식" in i:

            col_idx = 1
            write_ws.cell(row_idx, col_idx, i.split(" ")[0])
            col_idx += 1
            row_idx += 1
            write_ws.cell(row_idx, col_idx, i.split(" ")[1])


        elif "간편식" in i:
            col_idx = 1
            write_ws.cell(row_idx, col_idx, i.split(" ")[0])
            col_idx += 1
            row_idx += 1

            write_ws.cell(row_idx, col_idx, " ".join(i.split(" ")[1:]))



        else:
            write_ws.cell(row_idx, col_idx, i)
        row_idx += 1

    print("★☆★☆★☆★☆K4 Menu 입니다 ☆★☆★☆★☆★")
    row_idx = 2
    col_idx = 1

    write_ws = write_wb.active
    write_ws = write_wb.create_sheet('k4 menu')
    global k4_menu_table
    for i in k4_menu.split('\n'):
        if "중식" in i or "조식" in i or "석식" in i:
            write_ws['A1'] = i.split(" ")[0]
            k4_menu_table.append(i.split(" ")[0])
            for j in i.split(" "):
                if "한식" in j:
                    row_idx += 2
                    col_idx = 1
                    write_ws.cell(row_idx, col_idx, j)
                    k4_menu_table.append(j)
                    col_idx += 1
                    row_idx += 1
                elif which_menu not in j:
                    write_ws.cell(row_idx, col_idx, j)
                    k4_menu_table.append(j)
        elif "일품식" in i:
            col_idx = 1
            if ")" in i:
                write_ws.cell(row_idx, col_idx, i.split(")")[0] + ")")
                k4_menu_table.append(i.split(")")[0] + ")")
                col_idx += 1
                row_idx += 1
                write_ws.cell(row_idx, col_idx, i.split(")")[1])
                k4_menu_table.append(i.split(")")[1])

            else:
                write_ws.cell(row_idx, col_idx, i.split(" ")[0])
                k4_menu_table.append(i.split(" ")[0])
                col_idx += 1
                row_idx += 1
                write_ws.cell(row_idx, col_idx, i.split(" ")[1])
                k4_menu_table.append(i.split(" ")[1])
        elif "간편식" in i:
            col_idx = 1
            write_ws.cell(row_idx, col_idx, i.split(" ")[0])
            col_idx += 1
            row_idx += 1
            k4_menu_table.append(i.split(" ")[0])
            write_ws.cell(row_idx, col_idx, " ".join(i.split(" ")[1:]))

        else:
            write_ws.cell(row_idx, col_idx, i)
            k4_menu_table.append(i)
        row_idx += 1

    # write_ws['A1'] ="hi"

    # 행 단위로 추가
    # write_ws.append([1, 2, 3])

    # 셀 단위로 추가
    # write_ws.cell(5, 5, '5행5열')
    write_wb.save("Amkor " + which_menu + " Menu.xlsx")


def sending_mail():
    if not win32gui.FindWindow(None, "Microsoft Outlook"):  # 아웃룩이 실행되어있지 않은 경우
        os.startfile("outlook")  # 아웃룩 실행

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)

    mail.To = "rajeong.moon@amkor.co.kr;SuA.Lee@amkor.co.kr"

    file_location = f"C:\Myamkor\PycharmProjects\study\\"
    global which_menu

    attachment1 = file_location + "Amkor " + which_menu + " Menu.xlsx"
    attachment1 = os.path.abspath(attachment1)

    mail.Attachments.Add(Source=attachment1)

    hello_paragraph = "<p> 안녕하세요. Amkor " + which_menu + " 메뉴 리스트입니다.<br> </p>"

    print("html body")
    attachment2 = file_location + "image.jpg"
    mail.Attachments.Add(Source=attachment2)

    print("k4_menu_table")
    print(k4_menu_table)

    index_korea = 0
    index_one = 0
    index_easy = 0
    for i in range(len(k4_menu_table)):
        if k4_menu_table[i] == "한식":
            index_korea = i
        elif k4_menu_table[i] == "일품식":
            index_one = i
        elif k4_menu_table[i] == "간편식":
            index_easy = i
    k4_korean = []
    k4_one = []
    k4_easy = []
    for i in range(index_korea + 1, index_one):
        k4_korean.append(k4_menu_table[i])
    for i in range(index_one + 1, index_easy):
        k4_one.append(k4_menu_table[i])
    for i in range(index_easy + 1, len(k4_menu_table)):
        k4_easy.append(k4_menu_table[i])
    max_len = max(len(k4_korean), len(k4_easy), len(k4_one))
    while len(k4_one) != max_len:
        k4_one.append("")
    while len(k4_easy) != max_len:
        k4_easy.append("")
    while len(k4_korean) != max_len:
        k4_korean.append("")
    df_k4_menu_table = pd.DataFrame(
        {"한식": k4_korean, "일품식": k4_one, "간편식": k4_easy}
    )

    print(df_k4_menu_table)
    mail.HTMLBody = "<html><body> " + hello_paragraph + "<br> </body></html>" + \
                    df_k4_menu_table.to_html()
    # html table 만들기
    # <table><tr><th>     </th>   </tr>  <tr> <td>      </td>       </tr>        </table>

    # menu_info = "<p>" + " Menu LIST : <br> <table border=\"1\"> <tr>" \
    #                                            "<th>*</th> <th>*</th> <th>*</th> </tr>"
    #
    # global k5_menu,k4_menu
    #
    # td_str = ""
    # for i in k5_menu.split('\n'):
    #     menu_name= i.split(" ")
    #     for j in menu_name:
    #         print(j)
    #         td_str=td_str+"<td>"+j+"</td>"
    #     menu_info+="<tr>"+td_str+"</tr>"
    #
    # menu_info+="</table><br></p>"

    #     for j in range(0, len(lot_table_cell)):
    #         td_str = td_str + "<td>" + lot_table_cell[j] + "</td>"
    #     selected_lot_info = selected_lot_info + "<tr>" + td_str + "</tr>"
    # selected_lot_info = selected_lot_info + "</table><br>"

    #
    # 이미지 파일 이름에 띄어쓰기 들어가면 안됨.
    # mail.HTMLBody = "<html><body> <img src='cid:Damage Report.png'> <br> </body></html>"

    mail.Subject = "Amkor " + which_menu + " Menu"

    mail.FlagRequest = "Amkor " + which_menu + " Menu입니다"
    mail.Save()  # outlook 에서 draft 저장함.

    print("mail Display 시작. 완성된 메일 창 열기")

    mail.Display(False)  # outlook open//True면 기다리고 False면 파이썬 기다리지 않고 다음 스텝 진행함
    print("mail Display 완료+")
    # mail.Send()


if __name__ == '__main__':
    create_excel_of_lunch_menu()
    sending_mail()

    # try:
    #     os.rmdir("demofolder")
    # except Exception as e:
    #     shutil.rmtree("demofolder")

    # f = open("practice.txt", "r",encoding="UTF-8")
    # print(f.readline())
    # print(f.readline())
    # print(f.readlines())
    # print(f)
    # for x in f:
    #     print(x)
    #
    # f.close()
    #
    # try:
    #     f = open("demofile2.txt", "x")
    #     f.write("덮어쓰기~")
    #     f.close()
    # except Exception as e:
    #     print(e)
    #
    #
    #
    # if os.path.exists("demofile.txt"):
    #     os.remove("demofile.txt")
    # else:
    #     print("The file does not exist")
    #
    #     # open and read the file after the appending:
    # f = open("demofile2.txt", "r")
    # print(f.read())
    #
    # pass

    # make_excelfile(r"C:\SUA_PROJECT\pythonProject\excel test.xlsx")
    # read_excelfile(r"C:\SUA_PROJECT\pythonProject\device 별최적임계값 차트staticstics of threshold222.xlsx")

    # print(add(4,5))
    # print(subtract(4,5))

#
#     print_hi('PyCharm')
#     # a= 10;
#     # a=5;
#     # a= '';
#     # if a == '':
#     #     print("blank");
#     # elif a == 5 :
#     #     print("a 는 5");
#     # else :
#     #     print("a 는 10");
#     #
#     # for i in range(0,10):
#     #     print(i);
#     #
#     # arr= [5,3,2,1,2,4];
#     #
#     # for j in arr:
#     #     print(j);
#
#     dictval = dict();
#     dictval["apple"] = 5;
#     dictval["banana"]=3;
#     dictval["melon"]=8;
#
#     # for k in dictval.keys():
#     #     print(k)
#
#     # for j in dictval.values():
#     #     print(j)
#
#     # for l in dictval.keys():
#     #     print(dictval[l])
#
#     # try:
#     #     a=[4,5,2,4];
#     #     print(a[4]);
#     #
#     # except Exception as e:
#     #     print(e)
#     #
#
#     # a=[5,12,5,3,2];
#     # a=["b","c","eqsd","asda"]
#     # a.sort(reverse=True)
#
#     # dictVal=dict()
#     # dictVal["Jessy"]= "30"
#     # dictVal["Suzy"]= "28"
#     # dictVal["Minsu"]= "40"
#     # dictVal["Lily"]= "29"
#     #
#     # arr=[];
#     # for i in dictVal.values():
#     #     arr.append(i)
#     #
#     # arr.sort()
#     # val=[]
#     # for j in arr:
#     #     print(f"j:{j}")
#     #     for k in dictVal.keys():
#     #         print(f"key:{k}")
#     #         if (dictVal[k] == j):
#     #             val.append(k)
#
#
#
#     # print(a)
#     print(val)
# # See PyCharm help at https://www.jetbrains.com/help/pycharm/
#
#
