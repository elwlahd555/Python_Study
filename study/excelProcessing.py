from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

def read_excelfile(excel_path):
    load_wb = load_workbook(excel_path, data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['Sheet1']

    # 셀 주소로 값 출력
    print(load_ws['B2'].value)

    # 셀 좌표로 값 출력
    print(load_ws.cell(5, 2).value)
    a= load_ws["b2"].value


def make_excelfile(excel_path):
    wb = Workbook()
    sheet1 = wb.active
    # ws2 = wb.create_sheet(title="Pi")


    sheet1.title = "Device Level"
    blueGrayFill = PatternFill(fgColor="44546a", fill_type='solid')
    blueGrayLighterFill = PatternFill(fgColor="d6dce4", fill_type='solid')

    sheet1['A8'].fill = blueGrayFill
    sheet1['A8'].font = Font(name='맑은 고딕', size=11, bold=False, color='FFFFFF')

    sheet1.row_dimensions[1].height = 35

    sheet1.column_dimensions["A"].width = 5
    sheet1.column_dimensions["B"].width = 10
    sheet1.column_dimensions["C"].width = 16
    sheet1.column_dimensions["D"].width = 25
    sheet1.column_dimensions["E"].width = 20

    sheet1.freeze_panes = 'E10'

    sheet1['A9']= "안녕하세요"
    wb.save(excel_path)